"""Build overview xlsx workbooks from the benchmark cv_results CSVs.

One workbook per task (classification/ordinal), one sheet per feature source (cpg/markers/stacked), each cell holding
``mean±std`` over the CV folds. Written next to the CSVs under ``results/classifier/cv/``.

Sheet layout adapts to the CSV: when filter/reducer vary (the cpg benchmark) rows are filter x reducer and columns are
model x metric; when they do not (markers/stacked, always ``none``/``none``) rows are models and columns are metrics.

Usage:
    python -m python.helper.cv_results_to_xlsx                      # both tasks, all feature sources
    python -m python.helper.cv_results_to_xlsx --task ordinal
    python -m python.helper.cv_results_to_xlsx --csv <cv_results.csv> --out <summary.xlsx>   # single CSV
"""
import argparse
import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..config import CV_DIR

logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s][%(levelname)-5s][%(name)s:%(lineno)d] - %(message)s",
)
logger = logging.getLogger(__name__)

TASK_NAMES = ['classification', 'ordinal']
FEATURE_LABELS = {'cpg': 'CpG', 'markers': 'Markers', 'stacked': 'Stacked'}

FONT_NAME = "Arial"
HEADER1_FILL = PatternFill("solid", fgColor="4472C4")
HEADER2_FILL = PatternFill("solid", fgColor="5B9BD5")
BAND_FILL = PatternFill("solid", fgColor="F2F7FB")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
MAX_FILL = PatternFill("solid", fgColor="C6EFCE")

HEADER1_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
HEADER2_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name=FONT_NAME, size=10)
ROW_LABEL_FONT = Font(name=FONT_NAME, size=10, bold=True)
MAX_FONT = Font(name=FONT_NAME, size=10, bold=True, color="006100")
DASH_FONT = Font(name=FONT_NAME, size=10, color="999999")
CENTER = Alignment(horizontal="center", vertical="center")

HEADER_SIDE = Side(style="thin", color="2F5496")
DATA_SIDE = Side(style="thin", color="D9D9D9")
GROUP_SIDE = Side(style="medium", color="4472C4")
HEADER_BORDER = Border(left=HEADER_SIDE, right=HEADER_SIDE, top=HEADER_SIDE, bottom=HEADER_SIDE)
DATA_BORDER = Border(left=DATA_SIDE, right=DATA_SIDE, top=DATA_SIDE, bottom=DATA_SIDE)
GROUP_END_BORDER = Border(left=DATA_SIDE, right=DATA_SIDE, top=DATA_SIDE, bottom=GROUP_SIDE)

NON_METRIC = {"fold", "n_cpgs", "n_features"}
LOWER_IS_BETTER = {"mae", "mae_macro", "rmse", "mse", "loss"}
METRIC_LABELS = {
    "auroc": "AUROC",
    "bal_acc": "Bal. Acc.",
    "f1_macro": "F1",
    "mae": "MAE",
    "mae_macro": "MAE (macro)",
    "qwk": "QWK",
    "spearman_rho": "Spearman ρ",
    "rmse": "RMSE",
}
MODEL_LABELS = {
    "random_forest": "Random Forest",
    "xgboost": "XGBoost",
    "elasticnet": "ElasticNet",
    "logistic": "Logistic",
    "svm": "SVM",
    "svr": "SVR",
    "mlp": "MLP",
    "cnn": "CNN",
    "ridge": "Ridge",
    "lasso": "Lasso",
    "coral": "CORAL",
    "corn": "CORN",
    "ordinal_mlp": "Ordinal MLP",
    "mord_ridge": "MORD Ridge",
    "mord_logistic_at": "MORD LogisticAT",
    "ogboost": "OGBoost",
}


def fmt(mean: float, std: float) -> str:
    if pd.isna(mean):
        return "—"
    return f"{mean:.2f}±{std:.2f}"


def _aggregate(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str], list[str], str | None]:
    """Collapse per-fold rows into mean/std per configuration and decide the sheet layout.

    :param df: Per-fold cv_results table.
    :return: ``(agg, row_cols, metric_cols, col_col)`` where ``agg`` carries ``<metric>__mean``/``__std`` columns,
        ``row_cols`` are the columns forming the row index, and ``col_col`` is the column whose values form the top
        header level (``None`` when metrics alone make up the columns).
    """
    cols = df.columns.tolist()
    # Per-group MAE has its own plots; keeping it here would triple the column count for the ordinal sheets.
    ignore_cols = {c for c in cols if c.startswith("mae_group_") or c.startswith("mae_stage_")}
    metric_cols = [
        c for c in cols
        if c not in NON_METRIC and c not in ignore_cols and c != "fold" and pd.api.types.is_numeric_dtype(df[c])
    ]
    group_cols = [c for c in cols if c not in NON_METRIC and c not in ignore_cols and c not in metric_cols]
    if "model" not in group_cols:
        raise ValueError("Expected a 'model' column.")

    # skipna=False so any NaN in a fold propagates: a configuration with a NaN
    # in one of its folds is treated the same as a fully-empty one (shows "—").
    agg = df.groupby(group_cols, sort=False)[metric_cols].agg(
        [("mean", lambda s: s.mean(skipna=False)), ("std", lambda s: s.std(skipna=False))]
    )
    agg.columns = [f"{a}__{b}" for a, b in agg.columns]
    agg = agg.reset_index()

    # Constant group columns (filter/reducer are always none/none for markers/stacked) carry no information.
    varying = [c for c in group_cols if c != "model" and df[c].nunique() > 1]
    if varying:
        return agg, varying, metric_cols, "model"
    return agg, ["model"], metric_cols, None


def _render(ws, agg: pd.DataFrame, row_cols: list[str], metric_cols: list[str], col_col: str | None) -> None:
    """Write one aggregated table onto a worksheet, with headers, banding and best-per-metric highlighting."""
    n_row = len(row_cols)
    n_metric = len(metric_cols)
    col_groups = list(dict.fromkeys(agg[col_col].tolist())) if col_col else [None]

    row_keys = list(dict.fromkeys(agg[row_cols].itertuples(index=False, name=None)))
    lookup = {(tuple(r[c] for c in row_cols), r[col_col] if col_col else None): r for _, r in agg.iterrows()}

    # ---- header rows ----
    header_row1 = [c.capitalize() for c in row_cols] + [""] * (n_metric * len(col_groups))
    header_row2 = [""] * n_row
    for i, g in enumerate(col_groups):
        if g is not None:
            header_row1[n_row + n_metric * i] = MODEL_LABELS.get(g, g)
        for mc in metric_cols:
            header_row2.append(METRIC_LABELS.get(mc, mc))
    if col_col is None:
        # No top level: metric labels move up and span both header rows.
        header_row1[n_row:] = header_row2[n_row:]
        header_row2 = [""] * (n_row + n_metric)
    ws.append(header_row1)
    ws.append(header_row2)

    # row-label columns (and, without a top level, the metric columns) merge across both header rows
    merged_cols = n_row if col_col else n_row + n_metric
    for col_idx in range(1, merged_cols + 1):
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER1_FONT
        cell.fill = HEADER1_FILL
        cell.alignment = CENTER
        cell.border = HEADER_BORDER
        ws.cell(row=2, column=col_idx).border = HEADER_BORDER

    if col_col:
        # top header level (row 1): merge each group across its metric columns
        for i in range(len(col_groups)):
            start = n_row + i * n_metric + 1
            end = start + n_metric - 1
            if end > start:
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            for col_idx in range(start, end + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = HEADER1_FONT
                cell.fill = HEADER1_FILL
                cell.alignment = CENTER
                cell.border = HEADER_BORDER

        # metric sub-headers (row 2): lighter blue
        for col_idx in range(n_row + 1, n_row + n_metric * len(col_groups) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.font = HEADER2_FONT
            cell.fill = HEADER2_FILL
            cell.alignment = CENTER
            cell.border = HEADER_BORDER

    # ---- data rows ----
    # best per metric across the whole sheet (min for loss-like metrics, max otherwise), at display precision
    best_per_metric = {
        mc: round((agg[f"{mc}__mean"].min() if mc in LOWER_IS_BETTER else agg[f"{mc}__mean"].max()), 2)
        for mc in metric_cols
    }

    # With several row columns, band and separate by the first one (the filter); with a single one, alternate per row.
    grouped_rows = n_row > 1
    prev_first = None
    band = True
    for idx, key in enumerate(row_keys):
        if grouped_rows:
            if prev_first is not None and key[0] != prev_first:
                band = not band
            is_last_in_group = idx == len(row_keys) - 1 or row_keys[idx + 1][0] != key[0]
        else:
            band = not band
            is_last_in_group = idx == len(row_keys) - 1
        fill = BAND_FILL if band else WHITE_FILL

        row_vals: list = []
        is_max_flags: list = []
        for i, c in enumerate(row_cols):
            v = key[i]
            label = MODEL_LABELS.get(v, v) if c == "model" else v
            # repeated leading label (same filter as the row above) stays blank
            row_vals.append(None if grouped_rows and i == 0 and v == prev_first else label)
            is_max_flags.append(False)
        for g in col_groups:
            r = lookup.get((key, g))
            for mc in metric_cols:
                if r is None or pd.isna(r[f"{mc}__mean"]):
                    row_vals.append("—")
                    is_max_flags.append(None)  # dash
                else:
                    row_vals.append(fmt(r[f"{mc}__mean"], r[f"{mc}__std"]))
                    is_max_flags.append(bool(round(r[f"{mc}__mean"], 2) == best_per_metric[mc]))

        ws.append(row_vals)
        excel_row = ws.max_row
        border = GROUP_END_BORDER if is_last_in_group else DATA_BORDER
        for col_idx, flag in enumerate(is_max_flags, start=1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.fill = fill
            cell.border = border
            if col_idx <= n_row:
                cell.font = ROW_LABEL_FONT if col_idx == 1 else DATA_FONT
            else:
                cell.font = DATA_FONT
                cell.alignment = CENTER
            if flag is True:
                cell.fill = MAX_FILL
                cell.font = MAX_FONT
            elif flag is None:
                cell.font = DASH_FONT

        prev_first = key[0]

    # column widths: first row-label col wider for long filter/model names
    for col_idx in range(1, ws.max_column + 1):
        if col_idx == 1:
            width = 26
        elif col_idx <= n_row:
            width = 14
        else:
            width = 12
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = ws.cell(row=3, column=n_row + 1).coordinate


def write_workbook(sheets: dict[str, Path], out_path: Path) -> None:
    """Render one sheet per ``{title: cv_results.csv}`` entry into a single workbook.

    :param sheets: Sheet title -> per-fold cv_results CSV. Missing files are skipped with a warning.
    :param out_path: Destination xlsx.
    """
    wb = Workbook()
    wb.remove(wb.active)
    for title, csv_path in sheets.items():
        if not csv_path.exists():
            logger.warning('skipping %s: %s not found', title, csv_path)
            continue
        df = pd.read_csv(csv_path)
        agg, row_cols, metric_cols, col_col = _aggregate(df)
        _render(wb.create_sheet(title), agg, row_cols, metric_cols, col_col)
        logger.info('%s: %d configurations from %s', title, len(agg), csv_path.name)

    if not wb.sheetnames:
        raise FileNotFoundError(f'No input CSVs found for {out_path.name}.')
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    logger.info('wrote %s', out_path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('--task', choices=TASK_NAMES, nargs='*', default=TASK_NAMES,
                        help='Tasks to summarize (default: both). One workbook each.')
    parser.add_argument('--features', choices=list(FEATURE_LABELS), nargs='*', default=list(FEATURE_LABELS),
                        help='Feature sources to include as sheets (default: all).')
    parser.add_argument('--csv', type=Path, default=None,
                        help='Summarize a single cv_results CSV instead of the benchmark matrix.')
    parser.add_argument('--out', type=Path, default=None,
                        help='Output xlsx (with --csv) or output directory (default: results/classifier/cv).')
    args = parser.parse_args()

    if args.csv:
        write_workbook({'Results': args.csv}, args.out or args.csv.with_suffix('.xlsx'))
        return

    out_dir = args.out or CV_DIR
    for task in args.task:
        sheets = {FEATURE_LABELS[f]: CV_DIR / f'cv_results__{task}__{f}.csv' for f in args.features}
        write_workbook(sheets, out_dir / f'cv_results__{task}.xlsx')


if __name__ == "__main__":
    main()
